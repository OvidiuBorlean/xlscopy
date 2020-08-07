import openpyxl as xl 
from xlsconfig import *
import os
import pyexcel as p


inputWorkSheet = inputWorkSheet
outputWorkSheet = outputWorkSheet
#global readColumn1
#global writeColumn1

def ExcelActivities(inputXLS):
  
    wb1 = xl.load_workbook(inputXLS) 
    ws1 = wb1.active

    wb2 = xl.load_workbook(outputFile) 
    ws2 = wb2.active 
    mr = ws1.max_row 
    mc = ws1.max_column 
    for i in range (1, mr +1): 
              
            if 'readColumn1' in globals() and 'writeColumn1' in globals():
                c1 = ws1.cell(row = i, column = readColumn1) 
                ws2.cell(row = j, column = writeColumn1).value = c1.value
                
            if 'readColumn2' in globals() and 'writeColumn2' in globals():
                c2 = ws1.cell(row = i, column = readColumn2)
                ws2.cell(row = j, column = writeColumn2).value = c2.value

            if 'readColumn3' in globals() and 'writeColumn3' in globals():
                c3 = ws1.cell(row = i, column = readColumn3) 
                ws2.cell(row = i, column = writeColumn3).value = c3.value

            if 'readColumn4' in globals() and 'writeColumn4' in globals():    
                c4 = ws1.cell(row = i, column = readColumn4) 
                ws2.cell(row = i, column = writeColumn4).value = c4.value

            if 'readColumn5' in globals() and 'writeColumn5' in globals():
                c5 = ws1.cell(row = i, column = readColumn5) 
                ws2.cell(row = i, column = writeColumn5).value = c5.value

            if 'readColumn6' in globals() and 'writeColumn6' in globals():
                c6 = ws1.cell(row = i, column = readColumn6) 
                ws2.cell(row = i, column = writeColumn6).value = c6.value

            if 'readColumn7' in globals() and 'writeColumn7' in globals():
                c7 = ws1.cell(row = i, column = readColumn7) 
                ws2.cell(row = i, column = writeColumn7).value = c7.value

            if 'readColumn8' in globals() and 'writeColumn8' in globals():
                c8 = ws1.cell(row = i, column = readColumn8) 
                ws2.cell(row = i, column = writeColumn8).value = c8.value
         
            if 'readColumn9' in globals() and 'writeColumn9' in globals():
                c10 = ws1.cell(row = i, column = readColumn9) 
                ws2.cell(row = i, column = writeColumn9).value = c10.value
        
            if 'readColumn10' in globals() and 'writeColumn10' in globals():
                c10 = ws1.cell(row = i, column = readColumn10) 
                ws2.cell(row = i, column = writeColumn10).value = c10.value
                
            if 'readColumn11' in globals() and 'writeColumn11' in globals():
                c11 = ws1.cell(row = i, column = readColumn11) 
                ws2.cell(row = i, column = writeColumn11).value = c11.value
  
            if 'readColumn12' in globals() and 'writeColumn12' in globals():
                c12 = ws1.cell(row = i, column = readColumn12) 
                ws2.cell(row = i, column = writeColumn12).value = c12.value

            if 'readColumn13' in globals() and 'writeColumn13' in globals():
                c13 = ws1.cell(row = i, column = readColumn13) 
                ws2.cell(row = i, column = writeColumn13).value = c13.value
          
            if 'readColumn14' in globals() and 'writeColumn14' in globals():
                c14 = ws1.cell(row = i, column = readColumn14) 
                ws2.cell(row = i, column = writeColumn14).value = c14.value

            if 'readColumn15' in globals() and 'writeColumn15' in globals():
                c15 = ws1.cell(row = i, column = readColumn15) 
                ws2.cell(row = i, column = writeColumn15).value = c15.value
 
            if 'readColumn16' in globals() and 'writeColumn16' in globals():
                c16 = ws1.cell(row = i, column = readColumn16) 
                ws2.cell(row = i, column = writeColumn16).value = c16.value
            if 'readColumn17' in globals() and 'writeColumn17' in globals():
                c17 = ws1.cell(row = i, column = readColumn17) 
                ws2.cell(row = i, column = writeColumn17).value = c17.value 
            if 'readColumn18' in globals() and 'writeColumn18' in globals():
                c18 = ws1.cell(row = i, column = readColumn18) 
                ws2.cell(row = i, column = writeColumn18).value = c18.value
            if 'readColumn19' in globals() and 'writeColumn19' in globals():
                c19 = ws1.cell(row = i, column = readColumn19) 
                ws2.cell(row = i, column = writeColumn19).value = c19.value
            if 'readColumn20' in globals() and 'writeColumn20' in globals():
                c20 = ws1.cell(row = i, column = readColumn20) 
                ws2.cell(row = i, column = writeColumn20).value = c20.value
            if 'readColumn21' in globals() and 'writeColumn21' in globals():
                c21 = ws1.cell(row = i, column = readColumn21) 
                ws2.cell(row = i, column = writeColumn21).value = c21.value
            if 'readColumn22' in globals() and 'writeColumn22' in globals():
                c22 = ws1.cell(row = i, column = readColumn22) 
                ws2.cell(row = i, column = writeColumn22).value = c22.value      
            if 'readColumn23' in globals() and 'writeColumn23' in globals():
                c23 = ws1.cell(row = i, column = readColumn23) 
                ws2.cell(row = i, column = writeColumn23).value = c23.value
            if 'readColumn24' in globals() and 'writeColumn24' in globals():
                c24 = ws1.cell(row = i, column = readColumn24) 
                ws2.cell(row = i, column = writeColumn24).value = c24.value      
            if 'readColumn125' in globals() and 'writeColumn25' in globals():
                c25 = ws1.cell(row = i, column = readColumn25) 
                ws2.cell(row = i, column = writeColumn25).value = c25.value      
            if 'readColumn26' in globals() and 'writeColumn26' in globals():
                c26 = ws1.cell(row = i, column = readColumn26) 
                ws2.cell(row = i, column = writeColumn26).value = c26.value
            
            if 'readColumn127' in globals() and 'writeColumn27' in globals():
                c27 = ws1.cell(row = i, column = readColumn27) 
                ws2.cell(row = i, column = writeColumn27).value = c27.value
            
            if 'readColumn28' in globals() and 'writeColumn28' in globals():
                c28 = ws1.cell(row = i, column = readColumn28) 
                ws2.cell(row = i, column = writeColumn28).value = c28.value
            
            if 'readColumn29' in globals() and 'writeColumn29' in globals():
                c29 = ws1.cell(row = i, column = readColumn29) 
                ws2.cell(row = i, column = writeColumn29).value = c29.value      
            
            if 'readColumn30' in globals() and 'writeColumn30' in globals():
                c30 = ws1.cell(row = i, column = readColumn30) 
                ws2.cell(row = i, column = writeColumn30).value = c30.value
            
            if 'readColumn31' in globals() and 'writeColumn31' in globals():
                c31 = ws1.cell(row = i, column = readColumn31) 
                ws2.cell(row = i, column = writeColumn31).value = c31.value      
            
            if 'readColumn32' in globals() and 'writeColumn32' in globals():
                c32 = ws1.cell(row = i, column = readColumn32) 
                ws2.cell(row = i, column = writeColumn32).value = c32.value
  
    wb2.save(str(outputFile)) 

    
if __name__ == "__main__":

    fileList = os.listdir("./")
    for fileName in fileList:
        if fileName.endswith(".xls") or fileName.endswith(".XLS"):
           print(fileName)
           p.save_book_as(file_name = fileName, dest_file_name = fileName + ".xlsx")
           os.remove(fileName)
    for xlsxFile in fileList:
        if xlsxFile.endswith(".xlsx") or xlsxFile.endswith(".XLSX"):
           print(xlsxFile)
           ExcelActivities(xlsxFile)
           